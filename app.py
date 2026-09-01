code = r'''import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook

# ============================================================
# UTILITY FUNCTIONS
# ============================================================
def safe_write(ws, row_idx, col_idx, value):
    try:
        ws.cell(row=row_idx, column=col_idx).value = value
    except AttributeError:
        coord = ws.cell(row=row_idx, column=col_idx).coordinate
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                break


def get_column(df, possible_names):
    """Return the first matching column name found in df, else None."""
    for name in possible_names:
        if name in df.columns:
            return name
    return None


def parse_raw_table(raw):
    """
    Given a raw DataFrame (header=None) that may contain a blank leading
    column, a title row, and blank spacer rows, locate the real header
    row (the one containing a 'Date' cell) and return a clean DataFrame.
    """
    header_row_idx = None
    max_scan_rows = min(20, len(raw))

    for i in range(max_scan_rows):
        row_values = raw.iloc[i].astype(str).str.strip().str.lower()
        if (row_values == "date").any():
            header_row_idx = i
            break

    if header_row_idx is None:
        return None

    header = raw.iloc[header_row_idx].astype(str).str.strip()
    data = raw.iloc[header_row_idx + 1:].copy()
    data.columns = header

    valid_cols = [
        c for c in data.columns
        if str(c).strip() != "" and str(c).strip().lower() != "nan"
    ]
    data = data[valid_cols]
    data = data.dropna(how="all").reset_index(drop=True)

    return data


def load_timesheet_table(uploaded_file, sheet_name=None):
    """
    Robustly loads a single timesheet table from an uploaded CSV or Excel
    file, automatically detecting the real header row regardless of blank
    leading columns or title rows.
    """
    is_csv = uploaded_file.name.lower().endswith(".csv")

    if is_csv:
        raw = pd.read_csv(uploaded_file, header=None, dtype=str)
    else:
        target_sheet = sheet_name if sheet_name is not None else 0
        raw = pd.read_excel(uploaded_file, sheet_name=target_sheet, header=None, dtype=str)

    data = parse_raw_table(raw)

    if data is None:
        raise ValueError(
            "Could not find the timesheet header row (expected a column "
            "labeled 'Date'). Please check that the uploaded file is a "
            "valid timesheet export."
        )

    return data


def load_engineer_and_client_tables(uploaded_file):
    """
    Loads the Engineer-style and Client-style timesheet tables from a
    combined Excel workbook WITHOUT relying on specific tab names such
    as 'Engineer' or 'Client'.

    Instead, every sheet in the workbook is inspected, and sheets are
    matched based on the columns they actually contain:
      - Engineer-style sheets contain 'Travel OT' or 'Normal Time'.
      - Client-style sheets contain 'L.Trpt'.

    If the workbook only has one sheet, that same sheet is used for
    both the engineer-side and client-side figures.
    """
    if uploaded_file.name.lower().endswith(".csv"):
        client_df = load_timesheet_table(uploaded_file)
        eng_df = client_df.copy()
        return eng_df, client_df

    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    eng_df = None
    client_df = None

    for sheet in sheet_names:
        try:
            raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
            parsed = parse_raw_table(raw)
        except Exception:
            parsed = None

        if parsed is None:
            continue

        cols = set(parsed.columns)

        if eng_df is None and ({"Travel OT", "Normal Time"} & cols):
            eng_df = parsed

        if client_df is None and ("L.Trpt" in cols):
            client_df = parsed

    # If only one sheet exists, or no distinct match was found,
    # fall back to using the first parsable sheet for whichever
    # table is still missing.
    if eng_df is None or client_df is None:
        fallback_df = None
        for sheet in sheet_names:
            try:
                raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
                parsed = parse_raw_table(raw)
            except Exception:
                parsed = None

            if parsed is not None:
                fallback_df = parsed
                break

        if fallback_df is None:
            raise ValueError(
                "Could not find a usable timesheet table in the uploaded "
                "file. Please check that it contains a column labeled 'Date'."
            )

        if eng_df is None:
            eng_df = fallback_df
        if client_df is None:
            client_df = fallback_df

    return eng_df, client_df


def process_invoice_logic(
    eng_df, client_df, template_file,
    cust_name, inv_address, del_address, reference, cust_po,
    proj_no, svc_type, vessel_name, vessel_no, engineer_name,
    include_admin_fee, position, currency
):
    # Clean column names to remove accidental trailing spaces
    eng_df.columns = eng_df.columns.astype(str).str.strip()
    client_df.columns = client_df.columns.astype(str).str.strip()

    # --- PROCESS ENGINEER-SIDE HOURS ---
    if "Date" in eng_df.columns:
        eng_df = eng_df[eng_df["Date"].astype(str) != "Total"]

    travel_col = get_column(eng_df, ["Travel"])
    travel = pd.to_numeric(eng_df[travel_col], errors="coerce").fillna(0).sum() if travel_col else 0.0

    travel_ot_col = get_column(eng_df, ["Travel OT"])
    travel_ot = pd.to_numeric(eng_df[travel_ot_col], errors="coerce").fillna(0).sum() if travel_ot_col else 0.0

    travel_sum = travel + travel_ot

    nt_col = get_column(eng_df, ["Normal Time", "NT"])
    nt_sum = pd.to_numeric(eng_df[nt_col], errors="coerce").fillna(0).sum() if nt_col else 0.0

    ot_col = get_column(eng_df, ["OT"])
    ot_sum = pd.to_numeric(eng_df[ot_col], errors="coerce").fillna(0).sum() if ot_col else 0.0

    waiting_col = get_column(eng_df, ["Waiting Time", "Waiting time", "Waiting", "Waiting Hours"])
    waiting_sum = pd.to_numeric(eng_df[waiting_col], errors="coerce").fillna(0).sum() if waiting_col else 0.0

    prep_col = get_column(eng_df, ["Preparation"])
    prep_sum = pd.to_numeric(eng_df[prep_col], errors="coerce").fillna(0).sum() if prep_col else 0.0

    # --- PROCESS CLIENT-SIDE LOCAL TRANSPORT ---
    if "Date" in client_df.columns:
        client_df = client_df[client_df["Date"].astype(str) != "Total"]

    ltrpt_col = get_column(client_df, ["L.Trpt"])
    l_trpt_sum = pd.to_numeric(client_df[ltrpt_col], errors="coerce").fillna(0).sum() if ltrpt_col else 0.0

    # --- LOAD AND FILL INVOICE TEMPLATE ---
    if template_file.name.lower().endswith(".csv"):
        raise ValueError("The Invoice Template must be an Excel file (.xlsx) to preserve formulas and formatting.")

    wb = load_workbook(template_file)

    sheet_map = {"SG": "SG", "CN": "CN", "KR": "KR ", "EUR": "EUR", "USD": "USD"}
    target_sheet = sheet_map.get(currency, currency)

    if target_sheet not in wb.sheetnames:
        if currency in wb.sheetnames:
            target_sheet = currency
        else:
            raise ValueError(f"The uploaded template does not contain a tab for {currency}.")

    ws = wb[target_sheet]

    # Inject Customer Information
    safe_write(ws, 7, 3, cust_name)
    safe_write(ws, 8, 3, inv_address)
    safe_write(ws, 9, 3, del_address)
    safe_write(ws, 10, 3, reference)
    safe_write(ws, 11, 3, cust_po)
    safe_write(ws, 12, 3, proj_no)
    safe_write(ws, 13, 3, svc_type)
    safe_write(ws, 14, 3, vessel_name)
    safe_write(ws, 15, 3, vessel_no)

    # Determine Row Offset based on Engineer Role
    r_offset = 20
    if position == "Service Engineer":
        r_offset = 30
    elif position == "Senior Service Engineer":
        r_offset = 40
    elif position == "Specialist Service Engineer":
        r_offset = 50

    # Inject Hours into Invoice Table
    safe_write(ws, r_offset + 1, 4, travel_sum if travel_sum > 0 else "")
    safe_write(ws, r_offset + 2, 4, nt_sum if nt_sum > 0 else "")
    safe_write(ws, r_offset + 3, 4, ot_sum if ot_sum > 0 else "")
    safe_write(ws, r_offset + 4, 4, waiting_sum if waiting_sum > 0 else "")
    safe_write(ws, r_offset + 5, 4, prep_sum if prep_sum > 0 else "")

    expense_row = 59
    local_transport_row = None

    # Dynamically locate Expenses and Local Transport rows
    for row_idx in range(50, 75):
        col_b_val = str(ws.cell(row=row_idx, column=2).value).strip()
        col_c_val = str(ws.cell(row=row_idx, column=3).value).strip()

        if "Expenses" in col_b_val:
            expense_row = row_idx

        if "local transport" in col_c_val.lower() or "transportation" in col_c_val.lower():
            local_transport_row = row_idx
            break

    if not local_transport_row:
        local_transport_row = 63
        safe_write(ws, local_transport_row, 3, "Local Transport")

    safe_write(ws, expense_row + 2, 3, engineer_name)

    if l_trpt_sum > 0:
        safe_write(ws, local_transport_row, 4, l_trpt_sum)

    # --- DYNAMIC INVOICE TOTALS & TAX LOGIC ---
    if currency == "CN":
        if include_admin_fee == "No":
            safe_write(ws, 82, 7, "-")
        else:
            safe_write(ws, 82, 7, "=0.1*G67")
        safe_write(ws, 84, 7, "=0.06*(SUM(G28,G38,G48,G58,G67,G74,G82))")
        safe_write(ws, 86, 7, "=SUM(G28,G38,G48,G58,G67,G74,G77:G80,G82,G84)")

    elif currency == "KR":
        if include_admin_fee == "No":
            safe_write(ws, 82, 3, "-")
        else:
            safe_write(ws, 82, 3, "=0.1*(SUM(G63:G66))")
        safe_write(ws, 84, 3, "=SUM(G41:G47,G61:G63,C82,G31:G37, G21:G27, G51:G57)")

    elif currency == "SG":
        if include_admin_fee == "No":
            safe_write(ws, 78, 3, "-")
        else:
            safe_write(ws, 78, 3, "=SUM(G61:G62)*0.1")
        safe_write(ws, 80, 3, "=SUM(G41:G47,G61:G63,C78,G31:G37, G21:G27, G51:G57)")

    elif currency == "EUR":
        if include_admin_fee == "No":
            safe_write(ws, 82, 3, "-")
        else:
            safe_write(ws, 82, 3, "=0.1*(SUM(G61:G66))")
        safe_write(ws, 84, 3, "=SUM(G41:G47,G61:G66,C82,G31:G37, G21:G27,G51:G57,G70:G73,G77:G80)")

    elif currency == "USD":
        if include_admin_fee == "No":
            safe_write(ws, 82, 3, "-")
        else:
            safe_write(ws, 82, 3, "=0.1*(SUM(G63:G66))")
        safe_write(ws, 84, 3, "=SUM(G21:G27,G31:G37,G41:G47,G51:G57,G61:G66,G70:G73,G77:G80,C82)")

    # Export Final Invoice
    invoice_output = io.BytesIO()
    wb.save(invoice_output)
    invoice_output.seek(0)

    return invoice_output


# ============================================================
# STREAMLIT UI
# ============================================================
st.set_page_config(page_title="Invoice Generator", layout="wide")

st.title("Final Invoice Generation")
st.write("Select your timesheet format and generate the final invoice template.")

tab1, tab2 = st.tabs(["Single Timesheet Upload (Combined Excel)", "Sana Timesheet Upload (.csv or .xlsx)"])

# ------------------------------------------------------------
# TAB 1: SINGLE TIMESHEET UPLOAD
# ------------------------------------------------------------
with tab1:
    st.markdown("### 1. Upload Required Files")
    col1, col2 = st.columns(2)
    with col1:
        timesheet_excel_t1 = st.file_uploader("Upload Processed Timesheet (Excel)", type=["xlsx"], key="ts_upload_t1")
    with col2:
        template_excel_t1 = st.file_uploader("Upload Blank Invoice Template", type=["xlsx"], key="inv_upload_t1")


    st.markdown("### 2. Enter Information")
    c1_t1, c2_t1 = st.columns(2)
    with c1_t1:
        work_order_t1 = st.text_input("Work Order Number", value="NeedsConfirmation", key="wo_t1")
        cust_name_t1 = st.text_input("Customer name", key="cust_name_t1")
        inv_address_t1 = st.text_input("Invoicing address", key="inv_addr_t1")
        del_address_t1 = st.text_input("Delivery address", key="del_addr_t1")
        reference_t1 = st.text_input("Reference", key="ref_t1")
    with c2_t1:
        cust_po_t1 = st.text_input("Customer PO", key="po_t1")
        proj_no_t1 = st.text_input("Project No", key="proj_t1")
        svc_type_t1 = st.text_input("Service Type", key="svc_t1")
        vessel_name_t1 = st.text_input("Vessel Name", key="vessel_t1")
        vessel_no_t1 = st.text_input("Vessel No (if applicable)", key="vessel_no_t1")
        engineer_name_invoice_t1 = st.text_input("Engineer Name (For Expenses)", key="eng_name_t1")

    st.markdown("### 3. Service & Role Details")
    c3_t1, c4_t1, c5_t1 = st.columns(3)
    with c3_t1:
        currency_t1 = st.selectbox("Select Currency:", ["SG", "CN", "KR", "EUR", "USD"], key="curr_t1")
    with c4_t1:
        include_admin_fee_t1 = st.radio("Include 10% Admin Fee?", ["Yes", "No"], key="admin_fee_t1")
    with c5_t1:
        position_t1 = st.selectbox("Assign Hours to Position:", [
            "Service Technician", "Service Engineer", "Senior Service Engineer", "Specialist Service Engineer"
        ], key="pos_t1")

    if st.button("Generate Final Invoice", type="primary", key="btn_t1"):
        if not timesheet_excel_t1 or not template_excel_t1:
            st.error("Please upload the Processed Timesheet AND the Invoice Template.")
        else:
            try:
                eng_df, client_df = load_engineer_and_client_tables(timesheet_excel_t1)

                output = process_invoice_logic(
                    eng_df, client_df, template_excel_t1,
                    cust_name_t1, inv_address_t1, del_address_t1, reference_t1, cust_po_t1,
                    proj_no_t1, svc_type_t1, vessel_name_t1, vessel_no_t1, engineer_name_invoice_t1,
                    include_admin_fee_t1, position_t1, currency_t1
                )

                st.success("Invoice Generated Successfully")

                wo_num_t1 = work_order_t1.strip() or "NeedsConfirmation"

                st.download_button(
                    label="Download Final Invoice",
                    data=output,
                    file_name=f"Client_Invoice_{wo_num_t1}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_t1"
                )

            except KeyError as e:
                st.error(f"Missing expected column in timesheet: {str(e)}. Please check the uploaded file format.")
            except ValueError as e:
                st.error(f"Value error encountered: {str(e)}")
            except zipfile.BadZipFile:
                st.error("One of the uploaded files is not a valid Excel file or is corrupted.")
            except Exception as e:
                st.error(f"An unexpected error occurred while processing the invoice: {str(e)}")

# ------------------------------------------------------------
# TAB 2: STANDALONE TIMESHEET UPLOAD (CLIENT TIMESHEET ONLY)
# ------------------------------------------------------------
with tab2:
    st.markdown("### 1. Upload Required Files")

    col1_t2, col2_t2 = st.columns(2)

    with col1_t2:
        client_timesheet_t2 = st.file_uploader(
            "Upload Client Timesheet (Required)",
            type=["xlsx", "csv"],
            key="cli_upload_t2"
        )

    with col2_t2:
        template_excel_t2 = st.file_uploader(
            "Upload Blank Invoice Template",
            type=["xlsx"],
            key="inv_upload_t2"
        )

    st.markdown("### 2. Enter Information")

    c1_t2, c2_t2 = st.columns(2)

    with c1_t2:
        work_order_t2 = st.text_input("Work Order Number", value="NeedsConfirmation", key="wo_t2")
        cust_name_t2 = st.text_input("Customer name", key="cust_name_t2")
        inv_address_t2 = st.text_input("Invoicing address", key="inv_addr_t2")
        del_address_t2 = st.text_input("Delivery address", key="del_addr_t2")
        reference_t2 = st.text_input("Reference", key="ref_t2")

    with c2_t2:
        cust_po_t2 = st.text_input("Customer PO", key="po_t2")
        proj_no_t2 = st.text_input("Project No", key="proj_t2")
        svc_type_t2 = st.text_input("Service Type", key="svc_t2")
        vessel_name_t2 = st.text_input("Vessel Name", key="vessel_t2")
        vessel_no_t2 = st.text_input("Vessel No (if applicable)", key="vessel_no_t2")
        engineer_name_invoice_t2 = st.text_input("Engineer Name (For Expenses)", key="eng_name_t2")

    st.markdown("### 3. Service & Role Details")

    c3_t2, c4_t2, c5_t2 = st.columns(3)

    with c3_t2:
        currency_t2 = st.selectbox("Select Currency:", ["SG", "CN", "KR", "EUR", "USD"], key="curr_t2")

    with c4_t2:
        include_admin_fee_t2 = st.radio("Include 10% Admin Fee?", ["Yes", "No"], key="admin_fee_t2")

    with c5_t2:
        position_t2 = st.selectbox("Assign Hours to Position:", [
            "Service Technician", "Service Engineer", "Senior Service Engineer", "Specialist Service Engineer"
        ], key="pos_t2")

    if st.button("Generate Final Invoice", type="primary", key="btn_t2"):
        if not client_timesheet_t2 or not template_excel_t2:
            st.error("Please upload the Client Timesheet and the Invoice Template.")
        else:
            try:
                client_df = load_timesheet_table(client_timesheet_t2)
                eng_df = client_df.copy()

                expected_hour_columns = {
                    "Travel", "Travel OT", "Normal Time", "NT", "OT",
                    "Waiting Time", "Waiting time", "Preparation"
                }

                detected_hour_columns = set(client_df.columns) & expected_hour_columns

                if not detected_hour_columns:
                    raise ValueError(
                        "No timesheet hour columns were found in the Client Timesheet. "
                        "Expected columns such as Travel, Normal Time, NT, OT, "
                        "Waiting Time, or Preparation. Detected columns were: "
                        + ", ".join(str(c) for c in client_df.columns)
                    )

                output = process_invoice_logic(
                    eng_df, client_df, template_excel_t2,
                    cust_name_t2, inv_address_t2, del_address_t2, reference_t2, cust_po_t2,
                    proj_no_t2, svc_type_t2, vessel_name_t2, vessel_no_t2, engineer_name_invoice_t2,
                    include_admin_fee_t2, position_t2, currency_t2
                )

                st.success("Invoice Generated Successfully")

                wo_num_t2 = work_order_t2.strip() or "NeedsConfirmation"

                st.download_button(
                    label="Download Final Invoice",
                    data=output,
                    file_name=f"Client_Invoice_{wo_num_t2}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_t2"
                )

            except KeyError as e:
                st.error(f"Missing expected column in timesheet: {str(e)}. Please check the uploaded file format.")
            except ValueError as e:
                st.error(f"Value error encountered: {str(e)}")
            except zipfile.BadZipFile:
                st.error("One of the uploaded files is not a valid Excel file or is corrupted.")
            except Exception as e:
                st.error(f"An unexpected error occurred while processing the invoice: {str(e)}")
'''

with open("app.py", "w") as f:
    f.write(code)

print("File written, length:", len(code))
